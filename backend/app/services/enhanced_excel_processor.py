"""Enhanced Excel Processor Service with advanced pandas metadata extraction and data quality assessment."""

import hashlib
import logging
import re
import warnings
from collections import Counter
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats

logger = logging.getLogger(__name__)

# Suppress warnings for cleaner logs
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)


class EnhancedExcelProcessor:
    """Enhanced service for processing Excel files with advanced metadata extraction and data quality assessment."""

    def __init__(self, data_directory: str = "data/excel_files"):
        """Initialize the enhanced Excel processor.

        Args:
            data_directory: Directory containing Excel files
        """
        self.data_directory = Path(data_directory)
        self.data_directory.mkdir(parents=True, exist_ok=True)

        # Supported Excel file extensions
        self.supported_extensions = {".xlsx", ".xls", ".xlsm"}

        # Maximum file size in MB
        self.max_file_size_mb = 100

        # Cache for file hashes and metadata
        self._file_cache = {}

        # Pattern matchers for data types
        self.email_pattern = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b")
        self.phone_pattern = re.compile(
            r"(\+?\d{1,4}[\s-]?)?\(?\d{3}\)?[\s-]?\d{3}[\s-]?\d{4}|^\d{3}-\d{4}$"
        )
        self.url_pattern = re.compile(
            r"https?://(?:[-\w.])+(?:[:\d]+)?(?:/(?:[\w/_.])*(?:\?(?:[\w&=%.])*)?(?:#(?:[\w.])*)?)?"
        )
        self.currency_pattern = re.compile(r"[$€£¥]\s*\d+(?:[,.]?\d+)*|\d+(?:[,.]?\d+)*\s*[$€£¥]")

    def process_all_files(self) -> List[Dict[str, Any]]:
        """Process all Excel files in the data directory with enhanced capabilities.
        
        Returns:
            List of processed file data with additional enhanced metadata
        """
        results = []
        
        # DEBUG: Log directory information
        logger.info(f"[DEBUG] Looking for files in: {self.data_directory}")
        logger.info(f"[DEBUG] Directory exists: {self.data_directory.exists()}")
        logger.info(f"[DEBUG] Is directory: {self.data_directory.is_dir()}")
        logger.info(f"[DEBUG] Absolute path: {self.data_directory.absolute()}")
        
        # Find all Excel files
        excel_files = []
        for ext in self.supported_extensions:
            found = list(self.data_directory.glob(f"*{ext}"))
            logger.info(f"[DEBUG] Found {len(found)} files with extension {ext}")
            excel_files.extend(found)
        
        # DEBUG: List all files in directory
        if self.data_directory.exists():
            all_files = list(self.data_directory.iterdir())
            logger.info(f"[DEBUG] All files in directory: {[f.name for f in all_files]}")
        
        logger.info(f"Found {len(excel_files)} Excel files to process with enhanced capabilities")
        
        for file_path in excel_files:
            try:
                result = self.process_excel_file_enhanced(file_path)
                results.append(result)
                logger.info(f"Successfully processed {file_path.name} with enhanced features")
            except Exception as e:
                logger.error(f"Failed to process {file_path} with enhanced processor: {e}")
                continue
        
        logger.info(f"Enhanced processing completed for {len(results)}/{len(excel_files)} files")
        return results

    def get_statistics(self) -> Dict[str, Any]:
        """Get enhanced Excel processor statistics.
        
        Returns:
            Dictionary with enhanced processor statistics
        """
        try:
            # Get base statistics from file processor
            base_stats = self.get_file_statistics() if hasattr(self, 'get_file_statistics') else {}
            
            # Enhanced statistics
            excel_files = []
            for ext in self.supported_extensions:
                excel_files.extend(self.data_directory.glob(f"*{ext}"))
            
            processed_files = len([f for f in excel_files if f.exists()])
            cache_entries = len(getattr(self, '_file_cache', {}))
            
            enhanced_stats = {
                "processor_type": "enhanced",
                "total_excel_files": len(excel_files),
                "processed_files": processed_files,
                "cache_entries": cache_entries,
                "cache_memory_usage_mb": getattr(self, '_cache_memory_usage', 0) / (1024 * 1024),
                "supported_extensions": list(self.supported_extensions),
                "max_file_size_mb": self.max_file_size_mb,
                "parallel_processing_enabled": getattr(self, 'enable_parallel_processing', False),
                "data_directory": str(self.data_directory)
            }
            
            # Merge with base stats if available
            if base_stats:
                enhanced_stats.update(base_stats)
            
            return enhanced_stats
            
        except Exception as e:
            logger.error(f"Error getting enhanced Excel processor statistics: {e}")
            return {
                "processor_type": "enhanced",
                "total_excel_files": 0,
                "processed_files": 0,
                "cache_entries": 0,
                "error": str(e)
            }

    def get_file_hash(self, file_path: Path) -> str:
        """Generate MD5 hash for a file."""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def validate_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Validate Excel file before processing."""
        if not file_path.exists():
            raise ValueError(f"File does not exist: {file_path}")

        if file_path.suffix.lower() not in self.supported_extensions:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")

        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > self.max_file_size_mb:
            raise ValueError(f"File too large: {file_size_mb:.2f}MB > {self.max_file_size_mb}MB")

        return {
            "file_name": file_path.name,
            "file_size_mb": round(file_size_mb, 2),
            "extension": file_path.suffix.lower(),
            "last_modified": datetime.fromtimestamp(file_path.stat().st_mtime),
        }

    def detect_data_patterns(self, series: pd.Series) -> Dict[str, Any]:
        """Detect patterns in data series for enhanced metadata."""
        patterns = {
            "email_count": 0,
            "phone_count": 0,
            "url_count": 0,
            "currency_count": 0,
            "date_patterns": [],
            "numeric_patterns": {},
            "text_patterns": {},
        }

        if series.dtype == "object" or pd.api.types.is_string_dtype(series):
            string_data = series.astype(str).dropna()

            # Pattern matching
            patterns["email_count"] = sum(1 for x in string_data if self.email_pattern.search(x))
            patterns["phone_count"] = sum(1 for x in string_data if self.phone_pattern.search(x))
            patterns["url_count"] = sum(1 for x in string_data if self.url_pattern.search(x))
            patterns["currency_count"] = sum(
                1 for x in string_data if self.currency_pattern.search(x)
            )

            # Text patterns
            if len(string_data) > 0:
                lengths = [len(str(x)) for x in string_data if str(x) != "nan"]
                if lengths:
                    patterns["text_patterns"] = {
                        "avg_length": round(np.mean(lengths), 2),
                        "min_length": min(lengths),
                        "max_length": max(lengths),
                        "common_prefixes": self._find_common_prefixes(string_data),
                        "common_suffixes": self._find_common_suffixes(string_data),
                        "encoding_issues": self._detect_encoding_issues(string_data),
                    }

        elif pd.api.types.is_numeric_dtype(series):
            numeric_data = series.dropna()
            if len(numeric_data) > 1:
                patterns["numeric_patterns"] = {
                    "is_integer": all(float(x).is_integer() for x in numeric_data if pd.notna(x)),
                    "has_negative": any(x < 0 for x in numeric_data),
                    "has_zero": any(x == 0 for x in numeric_data),
                    "decimal_places": self._analyze_decimal_places(numeric_data),
                    "distribution_type": self._detect_distribution(numeric_data),
                    "outlier_detection": self._advanced_outlier_detection(numeric_data),
                    "seasonality": (
                        self._detect_seasonality(numeric_data) if len(numeric_data) >= 12 else None
                    ),
                }

        return patterns

    def _find_common_prefixes(self, string_data: pd.Series, min_length: int = 2) -> List[str]:
        """Find common prefixes in string data."""
        if len(string_data) < 2:
            return []

        prefixes = Counter()
        for text in string_data[:1000]:  # Limit for performance
            text = str(text)
            for i in range(min_length, min(len(text) + 1, 8)):
                prefixes[text[:i]] += 1

        threshold = max(2, len(string_data) * 0.1)
        return [prefix for prefix, count in prefixes.most_common(5) if count >= threshold]

    def _find_common_suffixes(self, string_data: pd.Series, min_length: int = 2) -> List[str]:
        """Find common suffixes in string data."""
        if len(string_data) < 2:
            return []

        suffixes = Counter()
        for text in string_data[:1000]:  # Limit for performance
            text = str(text)
            for i in range(min_length, min(len(text) + 1, 8)):
                suffixes[text[-i:]] += 1

        threshold = max(2, len(string_data) * 0.1)
        return [suffix for suffix, count in suffixes.most_common(5) if count >= threshold]

    def _detect_encoding_issues(self, string_data: pd.Series) -> Dict[str, Any]:
        """Detect potential encoding issues in text data."""
        encoding_issues = {"non_ascii_count": 0, "control_char_count": 0, "suspicious_chars": []}

        suspicious_chars = Counter()
        for text in string_data[:100]:  # Sample for performance
            text = str(text)
            encoding_issues["non_ascii_count"] += sum(1 for char in text if ord(char) > 127)
            encoding_issues["control_char_count"] += sum(
                1 for char in text if ord(char) < 32 and char not in ["\t", "\n", "\r"]
            )

            # Look for common encoding error patterns
            for char in text:
                if char in ["�", "\ufffd", "\x00"]:
                    suspicious_chars[char] += 1

        encoding_issues["suspicious_chars"] = dict(suspicious_chars.most_common(5))
        return encoding_issues

    def _analyze_decimal_places(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Analyze decimal places in numeric data."""
        decimal_counts = []
        for val in numeric_data:
            if pd.notna(val):
                if isinstance(val, float):
                    decimal_str = f"{val:.10f}".rstrip("0")
                    if "." in decimal_str:
                        decimal_places = len(decimal_str.split(".")[1])
                    else:
                        decimal_places = 0
                else:
                    decimal_places = 0
                decimal_counts.append(decimal_places)

        if decimal_counts:
            return {
                "max_decimal_places": max(decimal_counts),
                "avg_decimal_places": round(np.mean(decimal_counts), 2),
                "common_decimal_places": dict(Counter(decimal_counts).most_common(3)),
                "precision_consistency": len(set(decimal_counts)) <= 2,
            }
        return {}

    def _detect_distribution(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Detect the likely distribution type of numeric data."""
        if len(numeric_data) < 10:
            return {"type": "insufficient_data", "confidence": 0.0}

        try:
            data_array = np.array(numeric_data)

            # Basic statistics
            skewness = float(stats.skew(data_array))
            kurtosis = float(stats.kurtosis(data_array))

            # Distribution tests
            distribution_info = {
                "skewness": round(skewness, 4),
                "kurtosis": round(kurtosis, 4),
                "type": "unknown",
                "confidence": 0.0,
            }

            # Test for normality
            if len(data_array) >= 20:
                _, p_normal = stats.normaltest(data_array)
                if p_normal > 0.05:
                    distribution_info["type"] = "likely_normal"
                    distribution_info["confidence"] = round(p_normal, 4)

            # Basic distribution characteristics
            if distribution_info["type"] == "unknown":
                if abs(skewness) < 0.5:
                    distribution_info["type"] = "approximately_symmetric"
                elif skewness > 1:
                    distribution_info["type"] = "right_skewed"
                elif skewness < -1:
                    distribution_info["type"] = "left_skewed"

            return distribution_info

        except Exception as e:
            logger.warning(f"Distribution analysis failed: {e}")
            return {"type": "analysis_failed", "confidence": 0.0}

    def _advanced_outlier_detection(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Advanced outlier detection using multiple methods."""
        if len(numeric_data) < 4:
            return {"method": "insufficient_data", "outliers": [], "outlier_count": 0}

        data_array = np.array(numeric_data)
        outlier_info = {
            "iqr_outliers": [],
            "z_score_outliers": [],
            "modified_z_outliers": [],
            "total_outliers": 0,
            "outlier_percentage": 0.0,
        }

        try:
            # IQR method
            Q1 = np.percentile(data_array, 25)
            Q3 = np.percentile(data_array, 75)
            IQR = Q3 - Q1

            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            iqr_outliers = data_array[(data_array < lower_bound) | (data_array > upper_bound)]
            outlier_info["iqr_outliers"] = iqr_outliers.tolist()

            # Z-score method
            if len(data_array) > 1:
                z_scores = np.abs(stats.zscore(data_array))
                z_outliers = data_array[z_scores > 3]
                outlier_info["z_score_outliers"] = z_outliers.tolist()

                # Modified Z-score method
                median = np.median(data_array)
                mad = np.median(np.abs(data_array - median))
                if mad != 0:
                    modified_z_scores = 0.6745 * (data_array - median) / mad
                    modified_z_outliers = data_array[np.abs(modified_z_scores) > 3.5]
                    outlier_info["modified_z_outliers"] = modified_z_outliers.tolist()

            # Combined outlier count
            all_outliers = set(
                outlier_info["iqr_outliers"]
                + outlier_info["z_score_outliers"]
                + outlier_info["modified_z_outliers"]
            )

            outlier_info["total_outliers"] = len(all_outliers)
            outlier_info["outlier_percentage"] = round(
                (len(all_outliers) / len(data_array)) * 100, 2
            )

        except Exception as e:
            logger.warning(f"Outlier detection failed: {e}")

        return outlier_info

    def _detect_seasonality(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Detect potential seasonality in numeric time series data."""
        if len(numeric_data) < 12:
            return {"has_seasonality": False, "confidence": 0.0}

        try:
            # Simple autocorrelation test for common seasonal patterns
            data_array = np.array(numeric_data)

            # Test for monthly seasonality (lag 12)
            if len(data_array) >= 24:
                monthly_corr = np.corrcoef(data_array[:-12], data_array[12:])[0, 1]

                # Test for quarterly seasonality (lag 4)
                quarterly_corr = 0
                if len(data_array) >= 8:
                    quarterly_corr = np.corrcoef(data_array[:-4], data_array[4:])[0, 1]

                return {
                    "has_seasonality": abs(monthly_corr) > 0.5 or abs(quarterly_corr) > 0.5,
                    "monthly_correlation": (
                        round(monthly_corr, 4) if not np.isnan(monthly_corr) else 0
                    ),
                    "quarterly_correlation": (
                        round(quarterly_corr, 4) if not np.isnan(quarterly_corr) else 0
                    ),
                    "confidence": (
                        max(abs(monthly_corr), abs(quarterly_corr))
                        if not np.isnan(monthly_corr)
                        else 0
                    ),
                }

        except Exception:
            pass

        return {"has_seasonality": False, "confidence": 0.0}

    def assess_data_quality(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Comprehensive data quality assessment."""
        quality_metrics = {
            "completeness_score": 0.0,
            "consistency_score": 0.0,
            "validity_score": 0.0,
            "accuracy_score": 0.0,
            "overall_quality": "unknown",
            "issues": [],
            "recommendations": [],
            "column_quality": {},
            "data_profiling": {
                "row_count": len(df),
                "column_count": len(df.columns),
                "memory_usage_mb": round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2),
                "duplicate_rows": df.duplicated().sum(),
                "empty_rows": df.isnull().all(axis=1).sum(),
                "mixed_types": 0,
            },
        }

        if df.empty:
            return quality_metrics

        total_cells = df.size
        null_cells = df.isnull().sum().sum()

        # Completeness (how much data is present)
        completeness = 1 - (null_cells / total_cells) if total_cells > 0 else 0
        quality_metrics["completeness_score"] = round(completeness, 3)

        # Analyze each column
        consistency_scores = []
        validity_scores = []
        accuracy_scores = []

        for col in df.columns:
            col_quality = self._assess_column_quality(df[col], str(col))
            quality_metrics["column_quality"][str(col)] = col_quality

            consistency_scores.append(col_quality["consistency_score"])
            validity_scores.append(col_quality["validity_score"])
            accuracy_scores.append(col_quality["accuracy_score"])

            quality_metrics["issues"].extend(col_quality["issues"])
            quality_metrics["recommendations"].extend(col_quality["recommendations"])

            if col_quality["has_mixed_types"]:
                quality_metrics["data_profiling"]["mixed_types"] += 1

        # Calculate aggregate scores
        quality_metrics["consistency_score"] = (
            round(np.mean(consistency_scores), 3) if consistency_scores else 0
        )
        quality_metrics["validity_score"] = (
            round(np.mean(validity_scores), 3) if validity_scores else 0
        )
        quality_metrics["accuracy_score"] = (
            round(np.mean(accuracy_scores), 3) if accuracy_scores else 0
        )

        # Overall quality score
        overall = (
                          completeness
                          + quality_metrics["consistency_score"]
                          + quality_metrics["validity_score"]
                          + quality_metrics["accuracy_score"]
                  ) / 4

        quality_metrics["overall_quality"] = self._categorize_quality(overall)

        # Data profiling insights
        quality_metrics["data_profiling"]["duplicate_percentage"] = round(
            (quality_metrics["data_profiling"]["duplicate_rows"] / len(df)) * 100, 2
        )
        quality_metrics["data_profiling"]["completeness_percentage"] = round(completeness * 100, 2)

        return quality_metrics

    def _assess_column_quality(self, series: pd.Series, col_name: str) -> Dict[str, Any]:
        """Comprehensive quality assessment of a single column."""
        col_quality = {
            "consistency_score": 1.0,
            "validity_score": 1.0,
            "accuracy_score": 1.0,
            "issues": [],
            "recommendations": [],
            "has_mixed_types": False,
            "outlier_count": 0,
            "data_integrity": {
                "null_count": series.isnull().sum(),
                "unique_count": series.nunique(),
                "duplicate_count": series.duplicated().sum(),
            },
        }

        non_null_data = series.dropna()
        if len(non_null_data) == 0:
            col_quality["issues"].append(f"Column '{col_name}' is completely empty")
            col_quality["recommendations"].append(
                f"Remove column '{col_name}' or provide default values"
            )
            col_quality["consistency_score"] = 0.0
            col_quality["validity_score"] = 0.0
            col_quality["accuracy_score"] = 0.0
            return col_quality

        # Check for mixed data types
        if series.dtype == "object":
            types = set(type(x).__name__ for x in non_null_data if pd.notna(x))
            if len(types) > 1:
                col_quality["has_mixed_types"] = True
                col_quality["consistency_score"] *= 0.6
                col_quality["issues"].append(f"Column '{col_name}' has mixed data types: {types}")
                col_quality["recommendations"].append(
                    f"Standardize data types in column '{col_name}'"
                )

        # Numeric column analysis
        elif pd.api.types.is_numeric_dtype(series):
            outlier_info = self._advanced_outlier_detection(non_null_data)
            col_quality["outlier_count"] = outlier_info["total_outliers"]

            if outlier_info["outlier_percentage"] > 5:  # More than 5% outliers
                col_quality["validity_score"] *= 0.8
                col_quality["issues"].append(
                    f"Column '{col_name}' has {outlier_info['outlier_percentage']:.1f}% outliers"
                )
                col_quality["recommendations"].append(f"Review outliers in column '{col_name}'")

            # Check for unrealistic values (e.g., negative ages, impossible dates)
            if "age" in col_name.lower() and non_null_data.min() < 0:
                col_quality["accuracy_score"] *= 0.7
                col_quality["issues"].append(f"Column '{col_name}' has negative age values")

            if non_null_data.max() > 1e10:  # Very large numbers might indicate data errors
                col_quality["accuracy_score"] *= 0.9
                col_quality["issues"].append(f"Column '{col_name}' has unusually large values")

        # Check uniqueness patterns
        uniqueness_ratio = len(non_null_data.unique()) / len(non_null_data)

        if uniqueness_ratio < 0.1 and len(non_null_data) > 10:  # Less than 10% unique
            col_quality["consistency_score"] *= 0.9
            col_quality["issues"].append(
                f"Column '{col_name}' has very low diversity ({uniqueness_ratio:.1%} unique)"
            )

        # Check for potential ID columns with gaps
        if col_name.lower() in ["id", "index", "key"] and pd.api.types.is_numeric_dtype(series):
            if non_null_data.max() - non_null_data.min() + 1 != len(non_null_data.unique()):
                col_quality["consistency_score"] *= 0.8
                col_quality["issues"].append(f"ID column '{col_name}' appears to have gaps")

        return col_quality

    def _categorize_quality(self, score: float) -> str:
        """Categorize quality score."""
        if score >= 0.9:
            return "excellent"
        elif score >= 0.8:
            return "good"
        elif score >= 0.7:
            return "fair"
        elif score >= 0.6:
            return "poor"
        else:
            return "very_poor"

    def analyze_data_relationships(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze relationships between columns."""
        relationships = {
            "correlations": {},
            "potential_keys": [],
            "foreign_key_candidates": [],
            "functional_dependencies": [],
            "column_similarities": {},
        }

        if df.empty:
            return relationships

        # Correlation analysis for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 1:
            try:
                correlation_matrix = df[numeric_cols].corr()

                strong_correlations = {}
                for i, col1 in enumerate(numeric_cols):
                    for j, col2 in enumerate(numeric_cols):
                        if i < j:
                            corr_value = correlation_matrix.loc[col1, col2]
                            if abs(corr_value) > 0.7 and not pd.isna(corr_value):
                                strong_correlations[f"{col1}_vs_{col2}"] = {
                                    "correlation": round(corr_value, 4),
                                    "strength": "strong" if abs(corr_value) > 0.8 else "moderate",
                                }

                relationships["correlations"] = strong_correlations
            except Exception as e:
                logger.warning(f"Correlation analysis failed: {e}")

        # Identify potential primary keys
        for col in df.columns:
            non_null_data = df[col].dropna()
            if len(non_null_data) > 0:
                uniqueness_ratio = len(non_null_data.unique()) / len(df)
                if uniqueness_ratio > 0.95:
                    relationships["potential_keys"].append(
                        {
                            "column": str(col),
                            "uniqueness_ratio": round(uniqueness_ratio, 4),
                            "is_sequential": self._is_sequential(non_null_data),
                            "data_type": str(df[col].dtype),
                        }
                    )

        # Analyze column name similarities
        col_names = [str(col).lower() for col in df.columns]
        for i, col1 in enumerate(col_names):
            for j, col2 in enumerate(col_names):
                if i < j:
                    similarity = self._calculate_string_similarity(col1, col2)
                    if similarity > 0.7:
                        relationships["column_similarities"][
                            f"{df.columns[i]}_vs_{df.columns[j]}"
                        ] = {
                            "similarity_score": round(similarity, 3),
                            "potential_relationship": "naming_pattern",
                        }

        return relationships

    def _is_sequential(self, data: pd.Series) -> bool:
        """Check if numeric data is sequential."""
        if not pd.api.types.is_numeric_dtype(data):
            return False

        try:
            sorted_data = data.sort_values().reset_index(drop=True)
            differences = sorted_data.diff().dropna()
            return len(differences.unique()) == 1 and differences.iloc[0] == 1
        except Exception:
            return False

    def _calculate_string_similarity(self, s1: str, s2: str) -> float:
        """Calculate similarity between two strings using Jaccard similarity."""
        set1 = set(s1)
        set2 = set(s2)
        intersection = len(set1.intersection(set2))
        union = len(set1.union(set2))
        return intersection / union if union > 0 else 0

    def _detect_header(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Advanced header detection."""
        if df.empty:
            return {"has_header": False, "confidence": 0.0, "method": "empty_dataframe"}

        header_info = {
            "has_header": True,  # Default assumption
            "confidence": 0.5,
            "method": "default",
            "issues": [],
        }

        # Check for generic column names
        generic_pattern = re.compile(r"^Unnamed:\s*\d+$|^Column\s*\d+$", re.IGNORECASE)
        generic_count = sum(1 for col in df.columns if generic_pattern.match(str(col)))

        if generic_count > 0:
            header_info["has_header"] = False
            header_info["confidence"] = 0.9
            header_info["method"] = "generic_names_detected"
            header_info["issues"].append(f"Found {generic_count} generic column names")

        # Analyze first few rows for header patterns
        try:
            if len(df) > 1:
                first_row_types = [type(x).__name__ for x in df.iloc[0] if pd.notna(x)]
                second_row_types = [type(x).__name__ for x in df.iloc[1] if pd.notna(x)]

                # If first row is mostly strings and second row has mixed types, likely header
                first_row_strings = sum(1 for t in first_row_types if t == "str")
                if (
                        first_row_strings / len(first_row_types) > 0.8
                        and len(set(second_row_types)) > 1
                ):
                    header_info["has_header"] = True
                    header_info["confidence"] = 0.8
                    header_info["method"] = "type_analysis"

        except Exception as e:
            logger.warning(f"Header detection analysis failed: {e}")

        return header_info

    def extract_enhanced_metadata(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Extract comprehensive metadata from a DataFrame."""
        num_rows, num_cols = df.shape

        # Data quality assessment
        quality_assessment = self.assess_data_quality(df)

        # Data relationships
        relationships = self.analyze_data_relationships(df)

        # Header detection
        header_info = self._detect_header(df)

        # Enhanced column information
        columns_info = []
        for col in df.columns:
            col_data = df[col].dropna()
            col_info = {
                "name": str(col),
                "index": df.columns.get_loc(col),
                "dtype": str(df[col].dtype),
                "non_null_count": int(col_data.count()),
                "null_count": int(len(df) - col_data.count()),
                "null_percentage": round((len(df) - col_data.count()) / len(df) * 100, 2),
                "unique_count": int(col_data.nunique()),
                "uniqueness_ratio": (
                    round(col_data.nunique() / len(col_data), 4) if len(col_data) > 0 else 0
                ),
            }

            if len(col_data) > 0:
                # Data type specific analysis
                patterns = self.detect_data_patterns(df[col])

                if pd.api.types.is_numeric_dtype(df[col]):
                    col_info.update(
                        {
                            "data_type": "numeric",
                            "stats": {
                                "min": float(col_data.min()),
                                "max": float(col_data.max()),
                                "mean": round(float(col_data.mean()), 4),
                                "median": round(float(col_data.median()), 4),
                                "std": round(float(col_data.std()), 4),
                                "variance": round(float(col_data.var()), 4),
                                "quartiles": {
                                    "q1": round(float(col_data.quantile(0.25)), 4),
                                    "q3": round(float(col_data.quantile(0.75)), 4),
                                },
                                "skewness": (
                                    round(float(stats.skew(col_data)), 4)
                                    if len(col_data) > 1
                                    else 0
                                ),
                                "kurtosis": (
                                    round(float(stats.kurtosis(col_data)), 4)
                                    if len(col_data) > 1
                                    else 0
                                ),
                            },
                        }
                    )
                elif pd.api.types.is_datetime64_any_dtype(df[col]):
                    col_info.update(
                        {
                            "data_type": "datetime",
                            "stats": {
                                "min": str(col_data.min()),
                                "max": str(col_data.max()),
                                "date_range_days": (
                                    int((col_data.max() - col_data.min()).days)
                                    if len(col_data) > 1
                                    else 0
                                ),
                                "frequency_analysis": self._analyze_date_frequency(col_data),
                            },
                        }
                    )
                else:
                    col_info.update(
                        {
                            "data_type": "text",
                            "stats": {
                                "sample_values": col_data.unique()[:5].tolist(),
                                "most_frequent": dict(col_data.value_counts().head(3)),
                                "avg_length": round(col_data.astype(str).str.len().mean(), 2),
                                "encoding_info": patterns.get("text_patterns", {}).get(
                                    "encoding_issues", {}
                                ),
                            },
                        }
                    )

                col_info["patterns"] = patterns
                col_info["quality"] = quality_assessment["column_quality"].get(str(col), {})

            columns_info.append(col_info)

        # Sheet-level statistics
        memory_usage = df.memory_usage(deep=True).sum()

        return {
            "sheet_name": sheet_name,
            "num_rows": int(num_rows),
            "num_cols": int(num_cols),
            "columns": columns_info,
            "header_info": header_info,
            "data_preview": df.head(5).to_dict("records") if num_rows > 0 else [],
            "memory_usage_bytes": int(memory_usage),
            "data_quality": quality_assessment,
            "relationships": relationships,
            "sheet_statistics": {
                "total_cells": int(num_rows * num_cols),
                "non_empty_cells": int(df.count().sum()),
                "empty_cells": int(df.isnull().sum().sum()),
                "data_density": (
                    round(df.count().sum() / (num_rows * num_cols), 3)
                    if num_rows * num_cols > 0
                    else 0
                ),
                "numeric_columns": len(
                    [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]
                ),
                "text_columns": len([col for col in df.columns if df[col].dtype == "object"]),
                "datetime_columns": len(
                    [col for col in df.columns if pd.api.types.is_datetime64_any_dtype(df[col])]
                ),
                "duplicate_rows": int(df.duplicated().sum()),
                "data_types_distribution": dict(df.dtypes.value_counts()),
            },
        }

    def _analyze_date_frequency(self, date_series: pd.Series) -> Dict[str, Any]:
        """Analyze frequency patterns in datetime data."""
        if len(date_series) < 2:
            return {"frequency": "unknown", "intervals": []}

        try:
            sorted_dates = date_series.sort_values()
            intervals = sorted_dates.diff().dropna()

            # Most common interval
            interval_counts = intervals.value_counts()
            most_common_interval = interval_counts.index[0] if len(interval_counts) > 0 else None

            frequency_info = {
                "most_common_interval": str(most_common_interval),
                "interval_consistency": len(interval_counts) == 1,
                "total_intervals": len(intervals),
                "unique_intervals": len(interval_counts),
                "frequency": "irregular"  # Default value, will be overridden if pattern detected
            }

            # Detect common patterns
            if most_common_interval:
                days = most_common_interval.days
                if days == 1:
                    frequency_info["frequency"] = "daily"
                elif days == 7:
                    frequency_info["frequency"] = "weekly"
                elif 28 <= days <= 31:
                    frequency_info["frequency"] = "monthly"
                elif 90 <= days <= 92:
                    frequency_info["frequency"] = "quarterly"
                elif 365 <= days <= 366:
                    frequency_info["frequency"] = "yearly"
                else:
                    frequency_info["frequency"] = "irregular"

            return frequency_info

        except Exception:
            return {"frequency": "analysis_failed", "intervals": []}

    def create_enhanced_searchable_text(
            self, df: pd.DataFrame, sheet_name: str, metadata: Dict[str, Any]
    ) -> List[str]:
        """Create enhanced searchable text with quality and relationship insights."""
        text_chunks = []

        # Enhanced sheet summary
        quality_info = metadata.get("data_quality", {})
        summary = (
            f"Sheet: {sheet_name}\n"
            f"Dimensions: {metadata['num_rows']} rows × {metadata['num_cols']} columns\n"
            f"Data Quality: {quality_info.get('overall_quality', 'unknown')}\n"
            f"Completeness: {quality_info.get('completeness_score', 0):.1%}\n"
            f"Memory Usage: {metadata.get('memory_usage_bytes', 0) / 1024 / 1024:.1f} MB\n"
            f"Columns: {', '.join([col['name'] for col in metadata['columns']])}\n"
        )
        text_chunks.append(summary)

        # Enhanced column descriptions
        for col_info in metadata["columns"]:
            col_desc = (
                f"Column '{col_info['name']}' in sheet '{sheet_name}':\n"
                f"Type: {col_info['data_type']} ({col_info['dtype']})\n"
                f"Non-null values: {col_info['non_null_count']} ({100 - col_info['null_percentage']:.1f}%)\n"
                f"Unique values: {col_info['unique_count']} (uniqueness: {col_info['uniqueness_ratio']:.1%})\n"
            )

            # Add type-specific information
            if col_info["data_type"] == "numeric":
                stats = col_info.get("stats", {})
                patterns = col_info.get("patterns", {}).get("numeric_patterns", {})
                col_desc += (
                    f"Range: {stats.get('min', 'N/A')} to {stats.get('max', 'N/A')}\n"
                    f"Mean: {stats.get('mean', 'N/A')}, Median: {stats.get('median', 'N/A')}\n"
                    f"Distribution: {patterns.get('distribution_type', {}).get('type', 'unknown')}\n"
                )

                outlier_info = patterns.get("outlier_detection", {})
                if outlier_info.get("total_outliers", 0) > 0:
                    col_desc += f"Outliers detected: {outlier_info['total_outliers']} ({outlier_info['outlier_percentage']}%)\n"

            elif col_info["data_type"] == "text":
                stats = col_info.get("stats", {})
                patterns = col_info.get("patterns", {})
                col_desc += f"Average length: {stats.get('avg_length', 'N/A')} characters\n"
                col_desc += (
                    f"Sample values: {', '.join(map(str, stats.get('sample_values', [])))}\n"
                )

                # Add pattern information
                if patterns.get("email_count", 0) > 0:
                    col_desc += f"Contains {patterns['email_count']} email addresses\n"
                if patterns.get("phone_count", 0) > 0:
                    col_desc += f"Contains {patterns['phone_count']} phone numbers\n"
                if patterns.get("url_count", 0) > 0:
                    col_desc += f"Contains {patterns['url_count']} URLs\n"

            elif col_info["data_type"] == "datetime":
                stats = col_info.get("stats", {})
                col_desc += (
                    f"Date range: {stats.get('min', 'N/A')} to {stats.get('max', 'N/A')}\n"
                    f"Span: {stats.get('date_range_days', 'N/A')} days\n"
                )

                frequency_info = stats.get("frequency_analysis", {})
                if frequency_info.get("frequency") != "unknown":
                    col_desc += f"Frequency pattern: {frequency_info['frequency']}\n"

            # Add quality information
            quality = col_info.get("quality", {})
            if quality.get("issues"):
                col_desc += f"Quality issues: {'; '.join(quality['issues'][:2])}\n"

            text_chunks.append(col_desc)

        # Data quality summary
        if quality_info:
            quality_chunk = (
                f"Data Quality Analysis for sheet '{sheet_name}':\n"
                f"Overall Quality: {quality_info.get('overall_quality', 'unknown')}\n"
                f"Completeness: {quality_info.get('completeness_score', 0):.1%}\n"
                f"Consistency: {quality_info.get('consistency_score', 0):.1%}\n"
                f"Validity: {quality_info.get('validity_score', 0):.1%}\n"
            )

            if quality_info.get("issues"):
                quality_chunk += f"Key Issues: {'; '.join(quality_info['issues'][:3])}\n"

            text_chunks.append(quality_chunk)

        # Relationships summary
        relationships = metadata.get("relationships", {})
        if relationships.get("correlations") or relationships.get("potential_keys"):
            rel_chunk = f"Data Relationships in sheet '{sheet_name}':\n"

            if relationships.get("potential_keys"):
                key_names = [key["column"] for key in relationships["potential_keys"]]
                rel_chunk += f"Potential key columns: {', '.join(key_names)}\n"

            if relationships.get("correlations"):
                strong_corr = [
                    f"{k}: {v['correlation']:.2f}" for k, v in relationships["correlations"].items()
                ]
                rel_chunk += f"Strong correlations: {'; '.join(strong_corr[:3])}\n"

            text_chunks.append(rel_chunk)

        # Enhanced data samples with context
        chunk_size = 25  # Smaller chunks for better context
        for i in range(0, min(len(df), 200), chunk_size):  # Limit to first 200 rows
            chunk_df = df.iloc[i: i + chunk_size]

            chunk_text = f"Data sample from sheet '{sheet_name}' (rows {i + 1}-{i + len(chunk_df)}):\n"

            # Add statistical context for the chunk
            if pd.api.types.is_numeric_dtype(chunk_df.select_dtypes(include=[np.number])):
                numeric_summary = chunk_df.select_dtypes(include=[np.number]).describe()
                if not numeric_summary.empty:
                    chunk_text += (
                        f"Numeric summary for this chunk: {numeric_summary.loc['mean'].to_dict()}\n"
                    )

            # Add key-value pairs with enhanced context
            for idx, row in chunk_df.iterrows():
                row_items = []
                for col, value in row.items():
                    if pd.notna(value):
                        # Add contextual information
                        if isinstance(value, (int, float)) and abs(value) > 1000:
                            formatted_value = f"{value:,.0f}"
                        else:
                            formatted_value = str(value)
                        row_items.append(f"{col}: {formatted_value}")

                if row_items:
                    chunk_text += f"Row {idx}: " + " | ".join(row_items) + "\n"

            text_chunks.append(chunk_text)

        return text_chunks

    @lru_cache(maxsize=16)
    def process_excel_file_enhanced(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """Process Excel file with enhanced metadata extraction."""
        file_path = Path(file_path)

        try:
            file_info = self.validate_excel_file(file_path)
            file_hash = self.get_file_hash(file_path)

            logger.info(f"Processing Excel file with enhanced analysis: {file_path.name}")

            # Read all sheets
            try:
                if file_path.suffix.lower() in [".xlsx", ".xlsm"]:
                    excel_data = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
                else:
                    excel_data = pd.read_excel(file_path, sheet_name=None)
            except Exception as e:
                logger.warning(f"Failed to read with pandas, trying openpyxl: {e}")
                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                excel_data = {}
                for sheet_name in sheet_names:
                    try:
                        excel_data[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
                    except Exception as sheet_error:
                        logger.error(f"Failed to read sheet {sheet_name}: {sheet_error}")
                        continue

            # Process each sheet with enhanced analysis
            sheets_data = {}
            total_rows = 0
            total_columns = set()
            all_text_chunks = []
            file_quality_scores = []

            for sheet_name, df in excel_data.items():
                if df.empty:
                    continue

                # Clean DataFrame
                original_shape = df.shape
                df = df.dropna(how="all").dropna(axis=1, how="all")

                if df.empty:
                    continue

                # Enhanced metadata extraction
                metadata = self.extract_enhanced_metadata(df, sheet_name)

                # Enhanced searchable text
                text_chunks = self.create_enhanced_searchable_text(df, sheet_name, metadata)

                # Store enhanced sheet data
                sheets_data[sheet_name] = {
                    "metadata": metadata,
                    "text_chunks": text_chunks,
                    "data": df.to_dict("records")[:50],  # Store fewer rows but with richer metadata
                    "processing_info": {
                        "original_shape": original_shape,
                        "cleaned_shape": df.shape,
                        "cleaning_stats": {
                            "rows_removed": original_shape[0] - df.shape[0],
                            "cols_removed": original_shape[1] - df.shape[1],
                        },
                        "processing_time": datetime.now().isoformat(),
                    },
                }

                total_rows += metadata["num_rows"]
                total_columns.update([col["name"] for col in metadata["columns"]])
                all_text_chunks.extend(text_chunks)

                # Collect quality scores for file-level assessment
                quality_score = metadata.get("data_quality", {}).get("overall_quality", "unknown")
                if quality_score != "unknown":
                    quality_map = {"excellent": 5, "good": 4, "fair": 3, "poor": 2, "very_poor": 1}
                    file_quality_scores.append(quality_map.get(quality_score, 0))

            # File-level quality assessment
            file_quality = "unknown"
            if file_quality_scores:
                avg_score = np.mean(file_quality_scores)
                if avg_score >= 4.5:
                    file_quality = "excellent"
                elif avg_score >= 3.5:
                    file_quality = "good"
                elif avg_score >= 2.5:
                    file_quality = "fair"
                elif avg_score >= 1.5:
                    file_quality = "poor"
                else:
                    file_quality = "very_poor"

            result = {
                "file_name": file_info["file_name"],
                "file_hash": file_hash,
                "file_path": str(file_path),
                "file_size_mb": file_info["file_size_mb"],
                "last_modified": file_info["last_modified"],
                "extension": file_info["extension"],
                "total_sheets": len(sheets_data),
                "total_rows": total_rows,
                "total_columns": len(total_columns),
                "sheets": sheets_data,
                "all_text_chunks": all_text_chunks,
                "processed_at": datetime.now(),
                "processing_summary": {
                    "analysis_type": "enhanced",
                    "quality_assessment_enabled": True,
                    "relationship_analysis_enabled": True,
                    "pattern_detection_enabled": True,
                    "file_quality": file_quality,
                    "total_issues": sum(
                        len(sheet["metadata"].get("data_quality", {}).get("issues", []))
                        for sheet in sheets_data.values()
                    ),
                    "processing_version": "2.0",
                },
            }

            logger.info(
                f"Enhanced processing completed for {file_path.name}: "
                f"{len(sheets_data)} sheets, {total_rows} rows, "
                f"{len(all_text_chunks)} text chunks, quality: {file_quality}"
            )

            return result

        except Exception as e:
            logger.error(f"Error in enhanced processing of {file_path}: {e}")
            raise

    def get_enhanced_statistics(self) -> Dict[str, Any]:
        """Get comprehensive statistics about processed files."""
        try:
            all_files = []
            excel_files = []
            for ext in self.supported_extensions:
                excel_files.extend(self.data_directory.glob(f"*{ext}"))

            for file_path in excel_files:
                try:
                    result = self.process_excel_file_enhanced(file_path)
                    all_files.append(result)
                except Exception as e:
                    logger.error(f"Failed to process {file_path} for statistics: {e}")
                    continue

            if not all_files:
                return {
                    "total_files": 0,
                    "processing_summary": {"no_files_found": True},
                    "data_quality_summary": {"overall_quality": "no_data"},
                }

            # Enhanced statistics calculation
            stats = {
                "total_files": len(all_files),
                "total_sheets": sum(f["total_sheets"] for f in all_files),
                "total_rows": sum(f["total_rows"] for f in all_files),
                "total_columns": sum(f["total_columns"] for f in all_files),
                "total_text_chunks": sum(len(f["all_text_chunks"]) for f in all_files),
                "total_memory_usage_mb": round(
                    sum(
                        sheet["metadata"].get("memory_usage_bytes", 0)
                        for file_data in all_files
                        for sheet in file_data["sheets"].values()
                    )
                    / 1024
                    / 1024,
                    2,
                ),
                "file_types": dict(Counter(f.get("extension", "") for f in all_files)),
                "processing_summary": {
                    "enhanced_analysis_files": len(
                        [
                            f
                            for f in all_files
                            if f.get("processing_summary", {}).get("analysis_type") == "enhanced"
                        ]
                    ),
                    "average_file_quality": self._calculate_average_quality(all_files),
                    "total_data_issues": sum(
                        f.get("processing_summary", {}).get("total_issues", 0) for f in all_files
                    ),
                    "last_processed": max(
                        f.get("processed_at", datetime.min) for f in all_files
                    ).isoformat(),
                },
                "data_quality_distribution": self._analyze_quality_distribution(all_files),
                "column_type_distribution": self._analyze_column_types(all_files),
                "pattern_analysis": self._analyze_detected_patterns(all_files),
            }

            return stats

        except Exception as e:
            logger.error(f"Error generating enhanced statistics: {e}")
            return {"error": str(e), "total_files": 0}

    def _calculate_average_quality(self, all_files: List[Dict[str, Any]]) -> str:
        """Calculate average quality across all files."""
        quality_scores = []
        quality_map = {"excellent": 5, "good": 4, "fair": 3, "poor": 2, "very_poor": 1}

        for file_data in all_files:
            file_quality = file_data.get("processing_summary", {}).get("file_quality", "unknown")
            if file_quality in quality_map:
                quality_scores.append(quality_map[file_quality])

        if quality_scores:
            avg_score = np.mean(quality_scores)
            return self._categorize_quality(avg_score / 5)  # Normalize to 0-1

        return "unknown"

    def _analyze_quality_distribution(self, all_files: List[Dict[str, Any]]) -> Dict[str, int]:
        """Analyze distribution of quality scores."""
        quality_dist = Counter()

        for file_data in all_files:
            for sheet_data in file_data["sheets"].values():
                quality = (
                    sheet_data["metadata"].get("data_quality", {}).get("overall_quality", "unknown")
                )
                quality_dist[quality] += 1

        return dict(quality_dist)

    def _analyze_column_types(self, all_files: List[Dict[str, Any]]) -> Dict[str, int]:
        """Analyze distribution of column data types."""
        type_dist = Counter()

        for file_data in all_files:
            for sheet_data in file_data["sheets"].values():
                for col in sheet_data["metadata"].get("columns", []):
                    type_dist[col.get("data_type", "unknown")] += 1

        return dict(type_dist)

    def _analyze_detected_patterns(self, all_files: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze patterns detected across all files."""
        patterns_summary = {
            "email_columns": 0,
            "phone_columns": 0,
            "url_columns": 0,
            "currency_columns": 0,
            "outlier_affected_columns": 0,
            "mixed_type_columns": 0,
        }

        for file_data in all_files:
            for sheet_data in file_data["sheets"].values():
                for col in sheet_data["metadata"].get("columns", []):
                    patterns = col.get("patterns", {})

                    if patterns.get("email_count", 0) > 0:
                        patterns_summary["email_columns"] += 1
                    if patterns.get("phone_count", 0) > 0:
                        patterns_summary["phone_columns"] += 1
                    if patterns.get("url_count", 0) > 0:
                        patterns_summary["url_columns"] += 1
                    if patterns.get("currency_count", 0) > 0:
                        patterns_summary["currency_columns"] += 1

                    # Quality-related patterns
                    quality = col.get("quality", {})
                    if quality.get("outlier_count", 0) > 0:
                        patterns_summary["outlier_affected_columns"] += 1
                    if quality.get("has_mixed_types", False):
                        patterns_summary["mixed_type_columns"] += 1

        return patterns_summary
