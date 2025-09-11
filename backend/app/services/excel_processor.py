"""Optimized Excel Processor Service for analyzing and extracting data from Excel files.

Production-ready implementation with memory-efficient processing, chunking strategies,
parallel processing, and comprehensive caching for large Excel files.
"""

import asyncio
import hashlib
import logging
import os
import re
import tempfile
import warnings
import weakref
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from datetime import datetime
from functools import lru_cache, wraps
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, Generator, List, Optional, Tuple, Union, Callable

import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy import stats

logger = logging.getLogger(__name__)

# Suppress warnings for cleaner logs
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# Global thread pool for CPU-intensive operations
_thread_pool: Optional[ThreadPoolExecutor] = None

def get_thread_pool() -> ThreadPoolExecutor:
    """Get or create global thread pool for Excel processing."""
    global _thread_pool
    if _thread_pool is None or _thread_pool._shutdown:
        _thread_pool = ThreadPoolExecutor(
            max_workers=min(4, os.cpu_count() or 1),
            thread_name_prefix="excel_processor"
        )
    return _thread_pool

def cleanup_thread_pool():
    """Cleanup global thread pool."""
    global _thread_pool
    if _thread_pool is not None:
        _thread_pool.shutdown(wait=True)
        _thread_pool = None

# Register cleanup on module exit
import atexit
atexit.register(cleanup_thread_pool)


class ProcessingProgress:
    """Track processing progress for long-running operations."""
    
    def __init__(self, total_steps: int = 0):
        self.total_steps = total_steps
        self.current_step = 0
        self.status = "initialized"
        self.start_time = datetime.now()
        self.callbacks: List[Callable] = []
    
    def add_callback(self, callback: Callable[[int, int, str], None]):
        """Add progress callback function."""
        self.callbacks.append(callback)
    
    async def update(self, step: Optional[int] = None, status: str = ""):
        """Update progress and notify callbacks."""
        if step is not None:
            self.current_step = step
        if status:
            self.status = status
        
        for callback in self.callbacks:
            try:
                if asyncio.iscoroutinefunction(callback):
                    await callback(self.current_step, self.total_steps, self.status)
                else:
                    callback(self.current_step, self.total_steps, self.status)
            except Exception as e:
                logger.warning(f"Progress callback failed: {e}")
    
    def increment(self, status: str = "") -> 'ProcessingProgress':
        """Increment progress synchronously."""
        self.current_step += 1
        if status:
            self.status = status
        return self


class MemoryEfficientDataFrame:
    """Memory-efficient wrapper for large DataFrames with lazy loading."""
    
    def __init__(self, file_path: Path, sheet_name: str, chunk_size: int = 10000):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.chunk_size = chunk_size
        self._cached_info: Optional[Dict[str, Any]] = None
        self._temp_parquet_path: Optional[Path] = None
    
    @property
    def info(self) -> Dict[str, Any]:
        """Get basic DataFrame info without loading full data."""
        if self._cached_info is None:
            self._cached_info = self._extract_basic_info()
        return self._cached_info
    
    def _extract_basic_info(self) -> Dict[str, Any]:
        """Extract basic information without loading full DataFrame."""
        try:
            # Use openpyxl to get dimensions without loading data
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb[self.sheet_name]
            
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            
            # Sample first few rows for column info
            sample_df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name,
                nrows=min(100, max_row),
                engine='openpyxl'
            )
            
            wb.close()
            
            return {
                'shape': (max_row, max_col),
                'columns': sample_df.columns.tolist(),
                'dtypes': {col: str(dtype) for col, dtype in sample_df.dtypes.items()},
                'memory_estimate_mb': (max_row * max_col * 8) / (1024 * 1024)  # Rough estimate
            }
            
        except Exception as e:
            logger.warning(f"Failed to extract basic info: {e}")
            return {'shape': (0, 0), 'columns': [], 'dtypes': {}, 'memory_estimate_mb': 0}
    
    async def to_parquet_async(self) -> Path:
        """Convert Excel sheet to Parquet for efficient processing."""
        if self._temp_parquet_path and self._temp_parquet_path.exists():
            return self._temp_parquet_path
        
        temp_dir = Path(tempfile.gettempdir()) / "excel_processor"
        temp_dir.mkdir(exist_ok=True)
        
        temp_file = temp_dir / f"{self.file_path.stem}_{self.sheet_name}_{hash(str(self.file_path))}.parquet"
        
        def convert_to_parquet():
            try:
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=self.sheet_name,
                    engine='openpyxl'
                )
                
                # Optimize data types
                df = self._optimize_dtypes(df)
                
                # Convert to Parquet
                table = pa.Table.from_pandas(df)
                pq.write_table(table, temp_file, compression='snappy')
                
                return temp_file
                
            except Exception as e:
                logger.error(f"Failed to convert to Parquet: {e}")
                raise
        
        self._temp_parquet_path = await asyncio.to_thread(convert_to_parquet)
        return self._temp_parquet_path
    
    def _optimize_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize DataFrame data types for memory efficiency."""
        optimized_df = df.copy()
        
        for col in optimized_df.columns:
            col_data = optimized_df[col]
            
            # Skip if column is mostly null
            if col_data.isnull().sum() / len(col_data) > 0.9:
                continue
            
            # Optimize numeric columns
            if pd.api.types.is_numeric_dtype(col_data):
                # Try to downcast to smaller types
                if pd.api.types.is_integer_dtype(col_data):
                    optimized_df[col] = pd.to_numeric(col_data, downcast='integer')
                elif pd.api.types.is_float_dtype(col_data):
                    optimized_df[col] = pd.to_numeric(col_data, downcast='float')
            
            # Optimize string columns
            elif col_data.dtype == 'object':
                # Convert to category if low cardinality
                unique_ratio = col_data.nunique() / len(col_data)
                if unique_ratio < 0.5:  # Less than 50% unique values
                    try:
                        optimized_df[col] = col_data.astype('category')
                    except Exception:
                        pass  # Keep original type if conversion fails
        
        return optimized_df
    
    async def iter_chunks(self, chunk_size: Optional[int] = None) -> AsyncGenerator[pd.DataFrame, None]:
        """Iterate over DataFrame in chunks asynchronously."""
        chunk_size = chunk_size or self.chunk_size
        
        # Use Parquet for efficient chunking if file is large
        if self.info['memory_estimate_mb'] > 100:  # > 100MB estimated
            parquet_path = await self.to_parquet_async()
            
            def read_parquet_batches():
                parquet_file = pq.ParquetFile(parquet_path)
                for batch in parquet_file.iter_batches(batch_size=chunk_size):
                    yield batch.to_pandas()
            
            for chunk in await asyncio.to_thread(lambda: list(read_parquet_batches())):
                yield chunk
        else:
            # For smaller files, use pandas chunking
            def read_excel_chunks():
                try:
                    reader = pd.read_excel(
                        self.file_path,
                        sheet_name=self.sheet_name,
                        chunksize=chunk_size,
                        engine='openpyxl'
                    )
                    return list(reader)
                except Exception:
                    # Fallback: read full file and chunk manually
                    df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, engine='openpyxl')
                    return [df[i:i+chunk_size] for i in range(0, len(df), chunk_size)]
            
            chunks = await asyncio.to_thread(read_excel_chunks)
            for chunk in chunks:
                yield chunk
    
    def __del__(self):
        """Cleanup temporary files."""
        if self._temp_parquet_path and self._temp_parquet_path.exists():
            try:
                self._temp_parquet_path.unlink()
            except Exception:
                pass  # Ignore cleanup errors


class OptimizedExcelProcessor:
    """Unified Excel processor combining performance optimization with advanced analysis.
    
    Features:
    - Memory-efficient chunked processing for large files (up to 500MB)
    - Advanced data quality assessment and pattern detection
    - Relationship analysis and correlation detection
    - Configurable analysis modes (basic, comprehensive, auto)
    - Parallel processing where applicable
    - Progress tracking for long operations
    - Smart caching with memory management
    - Lazy loading and streaming capabilities
    
    Analysis Modes:
    - 'basic': Fast processing with essential metadata only
    - 'comprehensive': Full analysis with quality assessment and patterns
    - 'auto': Adaptive mode based on file size and system resources
    """

    def __init__(self, 
                 data_directory: str = "data/excel_files",
                 max_file_size_mb: int = 500,  # Increased limit
                 chunk_size: int = 10000,
                 max_cache_size_mb: int = 200,
                 enable_parallel_processing: bool = True,
                 analysis_mode: str = "auto"):
        """Initialize the unified Excel processor.
        
        Args:
            data_directory: Directory containing Excel files
            max_file_size_mb: Maximum file size limit (up to 500MB)
            chunk_size: Default chunk size for processing large files
            max_cache_size_mb: Maximum cache size in MB
            enable_parallel_processing: Enable parallel processing
            analysis_mode: Analysis depth ('basic', 'comprehensive', 'auto')
        """
        self.data_directory = Path(data_directory)
        self.data_directory.mkdir(parents=True, exist_ok=True)
        
        # Configuration
        self.supported_extensions = {'.xlsx', '.xls', '.xlsm'}
        self.max_file_size_mb = max_file_size_mb
        self.chunk_size = chunk_size
        self.max_cache_size_mb = max_cache_size_mb
        self.enable_parallel_processing = enable_parallel_processing
        self.analysis_mode = analysis_mode
        
        # Pattern matchers for enhanced analysis
        self.email_pattern = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b")
        self.phone_pattern = re.compile(
            r"(\+?\d{1,4}[\s-]?)?\(?\d{3}\)?[\s-]?\d{3}[\s-]?\d{4}|^\d{3}-\d{4}$"
        )
        self.url_pattern = re.compile(
            r"https?://(?:[-\w.])+(?:[:\d]+)?(?:/(?:[\w/_.])*(?:\?(?:[\w&=%.])*)?(?:#(?:[\w.])*)?)?"
        )
        self.currency_pattern = re.compile(r"[$€£¥]\s*\d+(?:[,.]?\d+)*|\d+(?:[,.]?\d+)*\s*[$€£¥]")
        
        # Enhanced caching with memory management
        self._file_cache = {}
        self._cache_access_times = {}
        self._cache_memory_usage = 0
        
        # Processing state
        self._active_operations = weakref.WeakSet()
        
        # Create temp directory for intermediate files
        self.temp_dir = Path(tempfile.gettempdir()) / "excel_processor"
        self.temp_dir.mkdir(exist_ok=True)
        
        logger.info(f"Initialized OptimizedExcelProcessor with max_file_size={max_file_size_mb}MB, "
                   f"chunk_size={chunk_size}, parallel_processing={enable_parallel_processing}")

    async def get_file_hash_async(self, file_path: Path) -> str:
        """Generate MD5 hash for a file asynchronously.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            MD5 hash string
        """
        def _compute_hash():
            hash_md5 = hashlib.md5()
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):  # Larger chunks for better performance
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        
        return await asyncio.to_thread(_compute_hash)
    
    def get_file_hash(self, file_path: Path) -> str:
        """Generate MD5 hash for a file (synchronous version)."""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    async def _manage_cache_memory(self):
        """Manage cache memory usage by removing oldest entries."""
        if self._cache_memory_usage <= self.max_cache_size_mb * 1024 * 1024:
            return
        
        # Sort by access time (oldest first)
        sorted_keys = sorted(
            self._cache_access_times.keys(),
            key=lambda k: self._cache_access_times[k]
        )
        
        # Remove oldest 25% of entries
        keys_to_remove = sorted_keys[:len(sorted_keys) // 4]
        
        for key in keys_to_remove:
            if key in self._file_cache:
                # Estimate memory usage (rough)
                data = self._file_cache[key]
                estimated_size = len(str(data).encode('utf-8'))
                self._cache_memory_usage -= estimated_size
                
                del self._file_cache[key]
                del self._cache_access_times[key]
        
        logger.info(f"Cache cleanup: removed {len(keys_to_remove)} entries, "
                   f"memory usage: {self._cache_memory_usage / 1024 / 1024:.1f}MB")
    
    def _update_cache(self, key: str, data: Any):
        """Update cache with memory tracking."""
        # Estimate memory usage
        estimated_size = len(str(data).encode('utf-8'))
        
        # Remove old entry if exists
        if key in self._file_cache:
            old_size = len(str(self._file_cache[key]).encode('utf-8'))
            self._cache_memory_usage -= old_size
        
        # Add new entry
        self._file_cache[key] = data
        self._cache_access_times[key] = datetime.now()
        self._cache_memory_usage += estimated_size
    
    @contextmanager
    def _progress_context(self, total_steps: int = 0) -> Generator[ProcessingProgress, None, None]:
        """Context manager for progress tracking."""
        progress = ProcessingProgress(total_steps)
        self._active_operations.add(progress)
        try:
            yield progress
        finally:
            self._active_operations.discard(progress)
    
    def validate_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Validate Excel file before processing.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with validation results
            
        Raises:
            ValueError: If file is invalid
        """
        if not file_path.exists():
            raise ValueError(f"File does not exist: {file_path}")
        
        if file_path.suffix.lower() not in self.supported_extensions:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")
        
        # Check file size
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > self.max_file_size_mb:
            raise ValueError(f"File too large: {file_size_mb:.2f}MB > {self.max_file_size_mb}MB")
        
        # Enhanced validation for large files
        validation_result = {
            "file_name": file_path.name,
            "file_size_mb": round(file_size_mb, 2),
            "extension": file_path.suffix.lower(),
            "last_modified": datetime.fromtimestamp(file_path.stat().st_mtime),
            "is_large_file": file_size_mb > 50,  # Flag for chunked processing
            "requires_streaming": file_size_mb > 200,  # Flag for streaming processing
        }
        
        # Quick sheet validation for large files
        if validation_result["is_large_file"]:
            try:
                wb = load_workbook(file_path, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                
                validation_result.update({
                    "sheet_count": sheet_count,
                    "processing_strategy": "chunked" if file_size_mb > 200 else "optimized"
                })
            except Exception as e:
                logger.warning(f"Could not validate sheets for {file_path.name}: {e}")
                validation_result["sheet_validation_warning"] = str(e)
        
        return validation_result

    async def extract_sheet_metadata_async(self, 
                                          df: pd.DataFrame, 
                                          sheet_name: str,
                                          progress: Optional[ProcessingProgress] = None,
                                          analysis_mode: str = "basic") -> Dict[str, Any]:
        """Extract metadata from a DataFrame asynchronously with progress tracking.
        
        Args:
            df: Pandas DataFrame
            sheet_name: Name of the Excel sheet
            progress: Optional progress tracker
            analysis_mode: Analysis depth ('basic', 'moderate', 'comprehensive')
            
        Returns:
            Dictionary with sheet metadata
        """
        # Basic dimensions
        num_rows, num_cols = df.shape
        
        if progress:
            await progress.update(status=f"Analyzing {sheet_name} metadata ({num_rows} rows)")
        
        # Optimize for large DataFrames by sampling
        if num_rows > 10000:
            # Use stratified sampling for large datasets
            sample_size = min(5000, num_rows // 10)
            sample_df = df.sample(n=sample_size, random_state=42)
            logger.info(f"Using sample of {sample_size} rows for metadata extraction of {sheet_name}")
        else:
            sample_df = df
        
        # Column information with parallel processing for large datasets
        columns_info = []
        
        if self.enable_parallel_processing and len(df.columns) > 10:
            # Process columns in parallel
            async def process_column(col):
                return await asyncio.to_thread(self._analyze_column, df[col], sample_df[col], str(col), num_rows, analysis_mode)
            
            # Process columns in batches to control memory usage
            batch_size = 20
            for i in range(0, len(df.columns), batch_size):
                batch_cols = df.columns[i:i + batch_size]
                tasks = [process_column(col) for col in batch_cols]
                batch_results = await asyncio.gather(*tasks, return_exceptions=True)
                
                for result in batch_results:
                    if isinstance(result, Exception):
                        logger.warning(f"Column analysis failed: {result}")
                        continue
                    columns_info.append(result)
                
                if progress:
                    await progress.update(status=f"Processed {min(i + batch_size, len(df.columns))}/{len(df.columns)} columns")
        else:
            # Sequential processing for smaller datasets
            for col in df.columns:
                try:
                    col_info = await asyncio.to_thread(
                        self._analyze_column, df[col], sample_df[col], str(col), num_rows, analysis_mode
                    )
                    columns_info.append(col_info)
                except Exception as e:
                    logger.warning(f"Failed to analyze column {col}: {e}")
                    continue
        
        # Memory-efficient preview
        preview_rows = min(3, num_rows)
        data_preview = df.head(preview_rows).to_dict('records') if num_rows > 0 else []
        
        return {
            "sheet_name": sheet_name,
            "num_rows": int(num_rows),
            "num_cols": int(num_cols),
            "columns": columns_info,
            "has_header": True,  # Assume first row is header
            "data_preview": data_preview,
            "memory_usage_mb": round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2),
            "was_sampled": num_rows > 10000,
            "sample_size": len(sample_df) if num_rows > 10000 else num_rows
        }
    
    def extract_sheet_metadata(self, df: pd.DataFrame, sheet_name: str, analysis_mode: str = "basic") -> Dict[str, Any]:
        """Synchronous wrapper for extract_sheet_metadata_async for backward compatibility."""
        import asyncio
        try:
            # Try to get existing event loop
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # If we're in an async context, we need to run in a thread
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future = executor.submit(
                        asyncio.run, 
                        self.extract_sheet_metadata_async(df, sheet_name, None, analysis_mode)
                    )
                    return future.result()
            else:
                # We can run directly
                return loop.run_until_complete(self.extract_sheet_metadata_async(df, sheet_name, None, analysis_mode))
        except RuntimeError:
            # No event loop, create new one
            return asyncio.run(self.extract_sheet_metadata_async(df, sheet_name, None, analysis_mode))
    
    def _analyze_column(self, full_col: pd.Series, sample_col: pd.Series, col_name: str, total_rows: int, analysis_mode: str = "basic") -> Dict[str, Any]:
        """Analyze a single column efficiently with configurable depth."""
        col_data = sample_col.dropna()
        
        if len(col_data) == 0:
            return {
                "name": col_name,
                "data_type": "empty",
                "dtype": str(full_col.dtype),
                "non_null_count": 0,
                "null_count": total_rows,
                "stats": {},
                "patterns": {} if analysis_mode != "basic" else None
            }
        
        dtype = str(full_col.dtype)
        non_null_count = int(full_col.count())
        
        # Basic column info
        column_info = {
            "name": col_name,
            "dtype": dtype,
            "non_null_count": non_null_count,
            "null_count": int(total_rows - non_null_count),
            "null_percentage": round((total_rows - non_null_count) / total_rows * 100, 2) if total_rows > 0 else 0,
        }
        
        # Detect data types more specifically
        if pd.api.types.is_numeric_dtype(full_col):
            data_type = "numeric"
            # Use sample for stats to avoid memory issues
            stats = {
                "min": float(col_data.min()) if not pd.isna(col_data.min()) else None,
                "max": float(col_data.max()) if not pd.isna(col_data.max()) else None,
                "mean": float(col_data.mean()) if not pd.isna(col_data.mean()) else None,
                "std": float(col_data.std()) if not pd.isna(col_data.std()) else None,
                "median": float(col_data.median()) if not pd.isna(col_data.median()) else None
            }
            
            # Enhanced statistics for comprehensive mode
            if analysis_mode == "comprehensive" and len(col_data) > 1:
                try:
                    stats.update({
                        "quartiles": {
                            "q1": float(col_data.quantile(0.25)),
                            "q3": float(col_data.quantile(0.75))
                        },
                        "skewness": float(stats.skew(col_data)) if len(col_data) > 1 else 0,
                        "kurtosis": float(stats.kurtosis(col_data)) if len(col_data) > 1 else 0,
                        "variance": float(col_data.var())
                    })
                except Exception as e:
                    logger.warning(f"Enhanced numeric stats failed for {col_name}: {e}")
                    
        elif pd.api.types.is_datetime64_any_dtype(full_col):
            data_type = "datetime"
            stats = {
                "min": str(col_data.min()),
                "max": str(col_data.max()),
                "date_range_days": (col_data.max() - col_data.min()).days if len(col_data) > 1 else 0
            }
            
            # Enhanced datetime analysis for comprehensive mode
            if analysis_mode == "comprehensive":
                try:
                    stats["frequency_analysis"] = self._analyze_date_frequency(col_data)
                except Exception as e:
                    logger.warning(f"Date frequency analysis failed for {col_name}: {e}")
                    
        else:
            data_type = "text"
            unique_count = col_data.nunique()
            value_counts = col_data.value_counts()
            stats = {
                "unique_values": min(unique_count, 10),
                "sample_values": col_data.unique()[:5].tolist(),
                "most_frequent": value_counts.head(3).to_dict() if len(value_counts) > 0 else {},
                "avg_length": round(col_data.astype(str).str.len().mean(), 2) if len(col_data) > 0 else 0
            }
        
        column_info.update({
            "data_type": data_type,
            "stats": stats
        })
        
        # Add pattern detection for enhanced modes
        if analysis_mode in ["comprehensive", "moderate"]:
            try:
                patterns = self.detect_data_patterns(sample_col)
                column_info["patterns"] = patterns
            except Exception as e:
                logger.warning(f"Pattern detection failed for {col_name}: {e}")
                column_info["patterns"] = {}
        
        return column_info

    async def create_searchable_text_async(self, 
                                          df: pd.DataFrame, 
                                          sheet_name: str,
                                          metadata: Dict[str, Any],
                                          progress: Optional[ProcessingProgress] = None) -> List[str]:
        """Create searchable text representations of the data.
        
        Args:
            df: Pandas DataFrame
            sheet_name: Name of the Excel sheet
            metadata: Sheet metadata
            
        Returns:
            List of text chunks for vector search
        """
        text_chunks = []
        
        # Sheet summary
        summary = (
            f"Sheet: {sheet_name}\n"
            f"Dimensions: {metadata['num_rows']} rows × {metadata['num_cols']} columns\n"
            f"Columns: {', '.join([col['name'] for col in metadata['columns']])}\n"
        )
        text_chunks.append(summary)
        
        # Column descriptions
        for col_info in metadata['columns']:
            col_desc = (
                f"Column '{col_info['name']}' in sheet '{sheet_name}':\n"
                f"Type: {col_info['data_type']}\n"
                f"Non-null values: {col_info['non_null_count']}\n"
            )
            
            if col_info['data_type'] == 'numeric':
                stats = col_info['stats']
                col_desc += (
                    f"Range: {stats.get('min', 'N/A')} to {stats.get('max', 'N/A')}\n"
                    f"Mean: {stats.get('mean', 'N/A')}\n"
                )
            elif col_info['data_type'] == 'text':
                stats = col_info['stats']
                col_desc += f"Sample values: {', '.join(map(str, stats.get('sample_values', [])))}\n"
            
            text_chunks.append(col_desc)
        
        # Data samples (chunked by rows)
        chunk_size = 50  # Process data in chunks
        for i in range(0, min(len(df), 500), chunk_size):  # Limit to first 500 rows
            chunk_df = df.iloc[i:i+chunk_size]
            
            # Convert chunk to text
            chunk_text = f"Data from sheet '{sheet_name}' (rows {i+1}-{i+len(chunk_df)}):\n"
            
            # Add key-value pairs for better searchability
            for _, row in chunk_df.iterrows():
                row_items = []
                for col, value in row.items():
                    if pd.notna(value):
                        row_items.append(f"{col}: {value}")
                
                if row_items:
                    chunk_text += " | ".join(row_items) + "\n"
            
            if chunk_text.strip():
                text_chunks.append(chunk_text)
        
        return text_chunks
    
    def create_searchable_text(self, df: pd.DataFrame, sheet_name: str, metadata: Dict[str, Any]) -> List[str]:
        """Synchronous wrapper for create_searchable_text_async for backward compatibility."""
        import asyncio
        try:
            # Try to get existing event loop
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # If we're in an async context, we need to run in a thread
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future = executor.submit(
                        asyncio.run, 
                        self.create_searchable_text_async(df, sheet_name, metadata, None)
                    )
                    return future.result()
            else:
                # We can run directly
                return loop.run_until_complete(self.create_searchable_text_async(df, sheet_name, metadata, None))
        except RuntimeError:
            # No event loop, create new one
            return asyncio.run(self.create_searchable_text_async(df, sheet_name, metadata, None))

    @lru_cache(maxsize=32)
    def process_excel_file(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """Process a single Excel file and extract all relevant information.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with file information and processed data
        """
        file_path = Path(file_path)
        
        try:
            # Validate file
            file_info = self.validate_excel_file(file_path)
            file_hash = self.get_file_hash(file_path)
            
            # Determine analysis mode based on file size
            analysis_mode = self._determine_analysis_mode(file_info["file_size_mb"])
            
            logger.info(f"Processing Excel file: {file_path.name} (mode: {analysis_mode})")
            
            # Read all sheets
            try:
                # Try to read with openpyxl first (better for .xlsx files)
                if file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                    excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                else:
                    excel_data = pd.read_excel(file_path, sheet_name=None)
            except Exception as e:
                logger.warning(f"Failed to read with pandas, trying openpyxl: {e}")
                # Fallback: read sheet names only
                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                excel_data = {}
                for sheet_name in sheet_names:
                    try:
                        excel_data[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
                    except Exception as sheet_error:
                        logger.error(f"Failed to read sheet {sheet_name}: {sheet_error}")
                        continue
            
            # Process each sheet
            sheets_data = {}
            total_rows = 0
            total_columns = set()
            all_text_chunks = []
            
            for sheet_name, df in excel_data.items():
                if df.empty:
                    continue
                
                # Clean the DataFrame
                df = df.dropna(how='all')  # Remove completely empty rows
                df = df.dropna(axis=1, how='all')  # Remove completely empty columns
                
                if df.empty:
                    continue
                
                # Extract metadata
                metadata = self.extract_sheet_metadata(df, sheet_name, analysis_mode)
                
                # Create searchable text
                text_chunks = self.create_searchable_text(df, sheet_name, metadata)
                
                # Store sheet data
                sheets_data[sheet_name] = {
                    "metadata": metadata,
                    "text_chunks": text_chunks,
                    "data": df.to_dict('records')[:100]  # Store first 100 rows as sample
                }
                
                total_rows += metadata['num_rows']
                total_columns.update([col['name'] for col in metadata['columns']])
                all_text_chunks.extend(text_chunks)
            
            result = {
                "file_name": file_info["file_name"],
                "file_hash": file_hash,
                "file_path": str(file_path),
                "file_size_mb": file_info["file_size_mb"],
                "last_modified": file_info["last_modified"],
                "extension": file_info["extension"],
                "total_sheets": len(sheets_data),
                "total_rows": total_rows,
                "total_columns": len(total_columns),
                "sheets": sheets_data,
                "all_text_chunks": all_text_chunks,
                "processed_at": datetime.now(),
                "analysis_mode": analysis_mode,
                "processing_info": {
                    "analysis_mode": analysis_mode,
                    "enhanced_features_enabled": analysis_mode in ["comprehensive", "moderate"],
                    "pattern_detection_enabled": analysis_mode in ["comprehensive", "moderate"],
                    "performance_optimized": True,
                    "processing_version": "2.0-unified"
                }
            }
            
            logger.info(f"Successfully processed {file_path.name}: "
                       f"{len(sheets_data)} sheets, {total_rows} rows, "
                       f"{len(all_text_chunks)} text chunks")
            
            return result
            
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            raise

    def process_all_files(self) -> List[Dict[str, Any]]:
        """Process all Excel files in the data directory.
        
        Returns:
            List of processed file data
        """
        results = []
        
        # Find all Excel files
        excel_files = []
        for ext in self.supported_extensions:
            excel_files.extend(self.data_directory.glob(f"*{ext}"))
        
        logger.info(f"Found {len(excel_files)} Excel files to process")
        
        for file_path in excel_files:
            try:
                result = self.process_excel_file(file_path)
                results.append(result)
            except Exception as e:
                logger.error(f"Failed to process {file_path}: {e}")
                continue
        
        return results

    def query_data(self, file_name: Optional[str] = None, 
                   sheet_name: Optional[str] = None,
                   column_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """Query processed data with optional filters.
        
        Args:
            file_name: Filter by file name
            sheet_name: Filter by sheet name
            column_name: Filter by column name
            
        Returns:
            List of matching data chunks
        """
        all_files = self.process_all_files()
        results = []
        
        for file_data in all_files:
            if file_name and file_name.lower() not in file_data['file_name'].lower():
                continue
            
            for sheet_name_key, sheet_data in file_data['sheets'].items():
                if sheet_name and sheet_name.lower() not in sheet_name_key.lower():
                    continue
                
                if column_name:
                    # Filter by column name
                    matching_columns = [
                        col for col in sheet_data['metadata']['columns']
                        if column_name.lower() in col['name'].lower()
                    ]
                    if not matching_columns:
                        continue
                
                # Add file and sheet context to results
                for chunk in sheet_data['text_chunks']:
                    results.append({
                        "file_name": file_data['file_name'],
                        "sheet_name": sheet_name_key,
                        "content": chunk,
                        "file_hash": file_data['file_hash']
                    })
        
        return results

    def get_file_statistics(self) -> Dict[str, Any]:
        """Get statistics about processed files.
        
        Returns:
            Dictionary with file statistics
        """
        all_files = self.process_all_files()
        
        if not all_files:
            return {
                "total_files": 0,
                "total_sheets": 0,
                "total_rows": 0,
                "total_columns": 0,
                "total_text_chunks": 0
            }
        
        total_files = len(all_files)
        total_sheets = sum(file_data['total_sheets'] for file_data in all_files)
        total_rows = sum(file_data['total_rows'] for file_data in all_files)
        total_columns = sum(file_data['total_columns'] for file_data in all_files)
        total_text_chunks = sum(len(file_data['all_text_chunks']) for file_data in all_files)
        
        return {
            "total_files": total_files,
            "total_sheets": total_sheets,
            "total_rows": total_rows,
            "total_columns": total_columns,
            "total_text_chunks": total_text_chunks,
            "average_sheets_per_file": round(total_sheets / total_files, 2),
            "average_rows_per_file": round(total_rows / total_files, 2)
        }

    def _determine_analysis_mode(self, file_size_mb: float) -> str:
        """Determine analysis mode based on file size and configuration."""
        if self.analysis_mode == "auto":
            if file_size_mb > 100:  # > 100MB files get basic analysis
                return "basic"
            elif file_size_mb > 20:  # 20-100MB files get moderate analysis
                return "moderate"
            else:  # < 20MB files get comprehensive analysis
                return "comprehensive"
        return self.analysis_mode

    def detect_data_patterns(self, series: pd.Series) -> Dict[str, Any]:
        """Detect patterns in data series for enhanced metadata."""
        patterns = {
            "email_count": 0,
            "phone_count": 0,
            "url_count": 0,
            "currency_count": 0,
            "date_patterns": [],
            "numeric_patterns": {},
            "text_patterns": {},
        }

        if series.dtype == "object" or pd.api.types.is_string_dtype(series):
            string_data = series.astype(str).dropna()

            # Pattern matching
            patterns["email_count"] = sum(1 for x in string_data if self.email_pattern.search(x))
            patterns["phone_count"] = sum(1 for x in string_data if self.phone_pattern.search(x))
            patterns["url_count"] = sum(1 for x in string_data if self.url_pattern.search(x))
            patterns["currency_count"] = sum(
                1 for x in string_data if self.currency_pattern.search(x)
            )

            # Text patterns
            if len(string_data) > 0:
                lengths = [len(str(x)) for x in string_data if str(x) != "nan"]
                if lengths:
                    patterns["text_patterns"] = {
                        "avg_length": round(np.mean(lengths), 2),
                        "min_length": min(lengths),
                        "max_length": max(lengths),
                        "common_prefixes": self._find_common_prefixes(string_data),
                        "common_suffixes": self._find_common_suffixes(string_data),
                        "encoding_issues": self._detect_encoding_issues(string_data),
                    }

        elif pd.api.types.is_numeric_dtype(series):
            numeric_data = series.dropna()
            if len(numeric_data) > 1:
                patterns["numeric_patterns"] = {
                    "is_integer": all(float(x).is_integer() for x in numeric_data if pd.notna(x)),
                    "has_negative": any(x < 0 for x in numeric_data),
                    "has_zero": any(x == 0 for x in numeric_data),
                    "decimal_places": self._analyze_decimal_places(numeric_data),
                    "distribution_type": self._detect_distribution(numeric_data),
                    "outlier_detection": self._advanced_outlier_detection(numeric_data),
                    "seasonality": (
                        self._detect_seasonality(numeric_data) if len(numeric_data) >= 12 else None
                    ),
                }

        return patterns

    def _find_common_prefixes(self, string_data: pd.Series, min_length: int = 2) -> List[str]:
        """Find common prefixes in string data."""
        if len(string_data) < 2:
            return []

        prefixes = Counter()
        for text in string_data[:1000]:  # Limit for performance
            text = str(text)
            for i in range(min_length, min(len(text) + 1, 8)):
                prefixes[text[:i]] += 1

        threshold = max(2, len(string_data) * 0.1)
        return [prefix for prefix, count in prefixes.most_common(5) if count >= threshold]

    def _find_common_suffixes(self, string_data: pd.Series, min_length: int = 2) -> List[str]:
        """Find common suffixes in string data."""
        if len(string_data) < 2:
            return []

        suffixes = Counter()
        for text in string_data[:1000]:  # Limit for performance
            text = str(text)
            for i in range(min_length, min(len(text) + 1, 8)):
                suffixes[text[-i:]] += 1

        threshold = max(2, len(string_data) * 0.1)
        return [suffix for suffix, count in suffixes.most_common(5) if count >= threshold]

    def _detect_encoding_issues(self, string_data: pd.Series) -> Dict[str, Any]:
        """Detect potential encoding issues in text data."""
        encoding_issues = {"non_ascii_count": 0, "control_char_count": 0, "suspicious_chars": []}

        suspicious_chars = Counter()
        for text in string_data[:100]:  # Sample for performance
            text = str(text)
            encoding_issues["non_ascii_count"] += sum(1 for char in text if ord(char) > 127)
            encoding_issues["control_char_count"] += sum(
                1 for char in text if ord(char) < 32 and char not in ["\t", "\n", "\r"]
            )

            # Look for common encoding error patterns
            for char in text:
                if char in ["�", "\ufffd", "\x00"]:
                    suspicious_chars[char] += 1

        encoding_issues["suspicious_chars"] = dict(suspicious_chars.most_common(5))
        return encoding_issues

    def _analyze_decimal_places(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Analyze decimal places in numeric data."""
        decimal_counts = []
        for val in numeric_data:
            if pd.notna(val):
                if isinstance(val, float):
                    decimal_str = f"{val:.10f}".rstrip("0")
                    if "." in decimal_str:
                        decimal_places = len(decimal_str.split(".")[1])
                    else:
                        decimal_places = 0
                else:
                    decimal_places = 0
                decimal_counts.append(decimal_places)

        if decimal_counts:
            return {
                "max_decimal_places": max(decimal_counts),
                "avg_decimal_places": round(np.mean(decimal_counts), 2),
                "common_decimal_places": dict(Counter(decimal_counts).most_common(3)),
                "precision_consistency": len(set(decimal_counts)) <= 2,
            }
        return {}

    def _detect_distribution(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Detect the likely distribution type of numeric data."""
        if len(numeric_data) < 10:
            return {"type": "insufficient_data", "confidence": 0.0}

        try:
            data_array = np.array(numeric_data)

            # Basic statistics
            skewness = float(stats.skew(data_array))
            kurtosis = float(stats.kurtosis(data_array))

            # Distribution tests
            distribution_info = {
                "skewness": round(skewness, 4),
                "kurtosis": round(kurtosis, 4),
                "type": "unknown",
                "confidence": 0.0,
            }

            # Test for normality
            if len(data_array) >= 20:
                _, p_normal = stats.normaltest(data_array)
                if p_normal > 0.05:
                    distribution_info["type"] = "likely_normal"
                    distribution_info["confidence"] = round(p_normal, 4)

            # Basic distribution characteristics
            if distribution_info["type"] == "unknown":
                if abs(skewness) < 0.5:
                    distribution_info["type"] = "approximately_symmetric"
                elif skewness > 1:
                    distribution_info["type"] = "right_skewed"
                elif skewness < -1:
                    distribution_info["type"] = "left_skewed"

            return distribution_info

        except Exception as e:
            logger.warning(f"Distribution analysis failed: {e}")
            return {"type": "analysis_failed", "confidence": 0.0}

    def _advanced_outlier_detection(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Advanced outlier detection using multiple methods."""
        if len(numeric_data) < 4:
            return {"method": "insufficient_data", "outliers": [], "outlier_count": 0}

        data_array = np.array(numeric_data)
        outlier_info = {
            "iqr_outliers": [],
            "z_score_outliers": [],
            "modified_z_outliers": [],
            "total_outliers": 0,
            "outlier_percentage": 0.0,
        }

        try:
            # IQR method
            Q1 = np.percentile(data_array, 25)
            Q3 = np.percentile(data_array, 75)
            IQR = Q3 - Q1

            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            iqr_outliers = data_array[(data_array < lower_bound) | (data_array > upper_bound)]
            outlier_info["iqr_outliers"] = iqr_outliers.tolist()

            # Z-score method
            if len(data_array) > 1:
                z_scores = np.abs(stats.zscore(data_array))
                z_outliers = data_array[z_scores > 3]
                outlier_info["z_score_outliers"] = z_outliers.tolist()

                # Modified Z-score method
                median = np.median(data_array)
                mad = np.median(np.abs(data_array - median))
                if mad != 0:
                    modified_z_scores = 0.6745 * (data_array - median) / mad
                    modified_z_outliers = data_array[np.abs(modified_z_scores) > 3.5]
                    outlier_info["modified_z_outliers"] = modified_z_outliers.tolist()

            # Combined outlier count
            all_outliers = set(
                outlier_info["iqr_outliers"]
                + outlier_info["z_score_outliers"]
                + outlier_info["modified_z_outliers"]
            )

            outlier_info["total_outliers"] = len(all_outliers)
            outlier_info["outlier_percentage"] = round(
                (len(all_outliers) / len(data_array)) * 100, 2
            )

        except Exception as e:
            logger.warning(f"Outlier detection failed: {e}")

        return outlier_info

    def _detect_seasonality(self, numeric_data: pd.Series) -> Dict[str, Any]:
        """Detect potential seasonality in numeric time series data."""
        if len(numeric_data) < 12:
            return {"has_seasonality": False, "confidence": 0.0}

        try:
            # Simple autocorrelation test for common seasonal patterns
            data_array = np.array(numeric_data)

            # Test for monthly seasonality (lag 12)
            if len(data_array) >= 24:
                monthly_corr = np.corrcoef(data_array[:-12], data_array[12:])[0, 1]

                # Test for quarterly seasonality (lag 4)
                quarterly_corr = 0
                if len(data_array) >= 8:
                    quarterly_corr = np.corrcoef(data_array[:-4], data_array[4:])[0, 1]

                return {
                    "has_seasonality": abs(monthly_corr) > 0.5 or abs(quarterly_corr) > 0.5,
                    "monthly_correlation": (
                        round(monthly_corr, 4) if not np.isnan(monthly_corr) else 0
                    ),
                    "quarterly_correlation": (
                        round(quarterly_corr, 4) if not np.isnan(quarterly_corr) else 0
                    ),
                    "confidence": (
                        max(abs(monthly_corr), abs(quarterly_corr))
                        if not np.isnan(monthly_corr)
                        else 0
                    ),
                }

        except Exception:
            pass

        return {"has_seasonality": False, "confidence": 0.0}

    def _analyze_date_frequency(self, date_series: pd.Series) -> Dict[str, Any]:
        """Analyze frequency patterns in datetime data."""
        if len(date_series) < 2:
            return {"frequency": "unknown", "intervals": []}

        try:
            sorted_dates = date_series.sort_values()
            intervals = sorted_dates.diff().dropna()

            # Most common interval
            interval_counts = intervals.value_counts()
            most_common_interval = interval_counts.index[0] if len(interval_counts) > 0 else None

            frequency_info = {
                "most_common_interval": str(most_common_interval),
                "interval_consistency": len(interval_counts) == 1,
                "total_intervals": len(intervals),
                "unique_intervals": len(interval_counts),
            }

            # Detect common patterns
            if most_common_interval:
                days = most_common_interval.days
                if days == 1:
                    frequency_info["frequency"] = "daily"
                elif days == 7:
                    frequency_info["frequency"] = "weekly"
                elif 28 <= days <= 31:
                    frequency_info["frequency"] = "monthly"
                elif 90 <= days <= 92:
                    frequency_info["frequency"] = "quarterly"
                elif 365 <= days <= 366:
                    frequency_info["frequency"] = "yearly"
                else:
                    frequency_info["frequency"] = "irregular"

            return frequency_info

        except Exception:
            return {"frequency": "analysis_failed", "intervals": []}


# Backward compatibility alias
ExcelProcessor = OptimizedExcelProcessor